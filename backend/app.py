from flask import Flask, request, jsonify, render_template, redirect, url_for, session, send_from_directory
from flask_cors import CORS
from datetime import datetime, timedelta, timezone
from db import get_connection
import hashlib
import os
import re
import secrets
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
CORS(app)
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))


# ─────────────────────────────────────────────
# MORTGAGE PROTECTION FUNNEL — CONFIG
# (protect-mortgage.com — see CLAUDE.md / .claude/rules)
# ─────────────────────────────────────────────

# ZIP allowlist — California only per CLAUDE.md §3. Add ranges here to add
# states; this is the config change, not a code change, and still requires
# written confirmation from John before enabling another state.
MP_ALLOWED_ZIP_RANGES = [(90001, 96162)]

MP_CODE_WORD_BLOCKLIST = {"password", "code", "codeword", "test", "none", "na"}
MP_CODE_WORD_PROFANITY_BLOCKLIST = {"fuck", "shit", "bitch", "asshole", "cunt", "nigger", "faggot"}


def mp_zip_allowed(zip_code):
    try:
        z = int(zip_code)
    except (TypeError, ValueError):
        return False
    return any(lo <= z <= hi for lo, hi in MP_ALLOWED_ZIP_RANGES)


def mp_validate_code_word(raw, first_name, last_name):
    """Returns an error message string, or None if the code word is valid."""
    word = (raw or "").strip()
    if not re.fullmatch(r"[A-Za-z]{3,20}", word):
        return "Code word must be 3-20 letters, no numbers or spaces."
    lower = word.lower()
    if lower == (first_name or "").strip().lower() or lower == (last_name or "").strip().lower():
        return "Please choose a code word other than your own name."
    if lower in MP_CODE_WORD_BLOCKLIST:
        return "That word is too common — please pick something more memorable."
    if lower in MP_CODE_WORD_PROFANITY_BLOCKLIST:
        return "Please choose a different word."
    return None


# ─────────────────────────────────────────────
# STATIC SITE
# ─────────────────────────────────────────────

SITE_ROOT = os.path.join(os.path.dirname(__file__), "..")
MAINTENANCE = os.getenv("MAINTENANCE_MODE", "").lower() == "true"

MAINTENANCE_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Coming Soon | Brown Financial Group</title>
  <link href="https://fonts.googleapis.com/css2?family=Raleway:wght@400;700;900&family=DM+Sans:wght@400;500&display=swap" rel="stylesheet"/>
  <style>
    *{box-sizing:border-box;margin:0;padding:0;}
    body{font-family:'DM Sans',sans-serif;background:#2C1A0E;color:#F5EFE6;min-height:100vh;display:flex;align-items:center;justify-content:center;text-align:center;padding:2rem;}
    .wrap{max-width:480px;}
    .logo{font-family:'Raleway',sans-serif;font-size:1.5rem;font-weight:900;color:#fff;letter-spacing:.02em;margin-bottom:2.5rem;}
    .logo span{color:#C08552;}
    h1{font-family:'Raleway',sans-serif;font-size:2.2rem;font-weight:700;line-height:1.2;margin-bottom:1rem;}
    p{font-size:1rem;line-height:1.7;color:#c8b89a;margin-bottom:2rem;}
    .divider{width:48px;height:3px;background:#7B3F1E;border-radius:2px;margin:0 auto 2rem;}
    a{color:#C08552;text-decoration:none;}
    a:hover{text-decoration:underline;}
    .contact{font-size:0.875rem;color:#a08060;}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="logo">Brown<span>Financial Group</span></div>
    <h1>Something great is coming.</h1>
    <div class="divider"></div>
    <p>We're putting the finishing touches on our new website. Check back soon — we can't wait to show you what we've built.</p>
    <p class="contact">In the meantime, reach us at<br/><a href="mailto:josh@thebrownfinancialgroup.com">josh@thebrownfinancialgroup.com</a></p>
  </div>
</body>
</html>"""

MAINTENANCE_PASS = {"/login", "/admin", "/logout", "/lead/"}

def is_protected(path):
    for prefix in MAINTENANCE_PASS:
        if path.startswith(prefix):
            return True
    return False

# Hosts that should land on the mortgage-protection intake instead of the
# main multi-product site. Once DNS points protect-mortgage.com at this
# service (see .claude/rules/infra.md), the domain root becomes the ad
# landing page — the two funnels stay on separate URLs but the same service.
MORTGAGE_PROTECTION_HOSTS = {"protect-mortgage.com", "www.protect-mortgage.com"}

def is_mortgage_protection_host():
    host = (request.host or "").lower().split(":")[0]
    return host in MORTGAGE_PROTECTION_HOSTS

@app.route("/")
def home():
    if MAINTENANCE and not is_protected("/"):
        return MAINTENANCE_HTML, 200
    if is_mortgage_protection_host():
        return send_from_directory(SITE_ROOT, "protect_mortgage.html")
    return send_from_directory(SITE_ROOT, "index.html")

@app.route("/thank-you")
def mortgage_protection_thank_you():
    return send_from_directory(SITE_ROOT, "mortgage_thank_you.html")

@app.route("/<path:filename>")
def static_site(filename):
    if MAINTENANCE and not is_protected("/" + filename):
        return MAINTENANCE_HTML, 200
    return send_from_directory(SITE_ROOT, filename)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def verify_password(stored_hash: str, password: str) -> bool:
    try:
        salt, hashed = stored_hash.split(":")
        return hashed == hashlib.sha256((salt + password).encode()).hexdigest()
    except Exception:
        return False


def get_current_agent():
    """Returns the logged-in agent row or None."""
    agent_id = session.get("agent_id")
    if not agent_id:
        return None
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT id, full_name, email, username, is_admin FROM agents WHERE id = %s AND is_active = TRUE",
            (agent_id,)
        )
        agent = cur.fetchone()
        cur.close()
        conn.close()
        return agent
    except Exception:
        return None


def send_lead_notification(data: dict):
    """
    Sends a new lead alert email to all agents with notify_on_lead = TRUE.
    Uses SendGrid API — works on Render free tier.
    Failures are silent so a mail issue never blocks a lead from being saved.
    """
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail

    api_key = os.getenv("SENDGRID_API_KEY")
    sender  = os.getenv("MAIL_SENDER")

    if not api_key or not sender:
        return

    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT email, full_name FROM agents WHERE notify_on_lead = TRUE AND is_active = TRUE"
        )
        recipients = cur.fetchall()
        cur.close()
        conn.close()
    except Exception:
        return

    if not recipients:
        return

    PRODUCT_LABELS = {
        'mortgage-protection': 'Mortgage Protection',
        'final-expense':       'Final Expense',
        'term-life':           'Term Life',
        'whole-life':          'Whole Life',
        'iul':                 'IUL / Wealth Building',
        'living-benefits':     'Living Benefits',
    }

    first        = data.get("first_name", "")
    last         = data.get("last_name", "")
    product      = PRODUCT_LABELS.get(data.get("product_type", ""), data.get("product_type", "Unknown"))
    mobile       = data.get("mobile_phone", "—")
    home_phone   = data.get("home_phone", "—") or "—"
    email        = data.get("email", "—") or "—"
    city         = data.get("city", "—") or "—"
    state        = data.get("state", "—") or "—"
    has_ben      = (data.get("has_beneficiary") or "—").capitalize()
    ben_rel      = data.get("beneficiary_relationship", "—") or "—"
    reason       = data.get("reason", "—") or "—"
    contact_pref = "Call Me" if data.get("contact_preference") == "call_me" else "Book Appointment"
    submitted    = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    subject = f"New Lead: {first} {last} — {product}"

    html_body = f"""
    <div style="font-family:'DM Sans',Arial,sans-serif;max-width:560px;margin:0 auto;background:#ffffff;border:1px solid #e8d5c0;border-radius:8px;overflow:hidden">
      <div style="background:#4B2E2B;padding:1.25rem 1.5rem">
        <p style="font-family:Arial,sans-serif;font-size:1.15rem;font-weight:800;color:#ffffff;margin:0">
          Brown<span style="color:#C08552">Financial Group</span>
        </p>
        <p style="font-size:0.75rem;color:#c8a882;margin:0.2rem 0 0;letter-spacing:0.08em;text-transform:uppercase">New Lead Notification</p>
      </div>
      <div style="padding:1.5rem">
        <h2 style="font-size:1.1rem;color:#4B2E2B;margin:0 0 0.25rem">{first} {last}</h2>
        <p style="font-size:0.85rem;color:#8C5A3C;margin:0 0 1.5rem;font-weight:600">{product}</p>

        <table style="width:100%;border-collapse:collapse;font-size:0.875rem">
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;width:160px;font-weight:600">Mobile</td>
            <td style="padding:0.6rem 0;color:#2C1810">{mobile}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Home Phone</td>
            <td style="padding:0.6rem 0;color:#2C1810">{home_phone}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Email</td>
            <td style="padding:0.6rem 0;color:#2C1810">{email}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">City</td>
            <td style="padding:0.6rem 0;color:#2C1810">{city}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">State</td>
            <td style="padding:0.6rem 0;color:#2C1810">{state}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Product Type</td>
            <td style="padding:0.6rem 0;color:#2C1810">{product}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Has Beneficiary</td>
            <td style="padding:0.6rem 0;color:#2C1810">{has_ben}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Beneficiary Rel.</td>
            <td style="padding:0.6rem 0;color:#2C1810">{ben_rel}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Reason</td>
            <td style="padding:0.6rem 0;color:#2C1810">{reason}</td>
          </tr>
          <tr style="border-bottom:1px solid #f0e4d4">
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Contact Pref</td>
            <td style="padding:0.6rem 0;color:#2C1810">{contact_pref}</td>
          </tr>
          <tr>
            <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Submitted</td>
            <td style="padding:0.6rem 0;color:#2C1810">{submitted}</td>
          </tr>
        </table>

        <div style="margin-top:1.5rem;padding-top:1.25rem;border-top:1px solid #f0e4d4">
          <a href="https://thebrownfinancialgroup.com/admin"
             style="display:inline-block;background:#C08552;color:#ffffff;font-weight:700;font-size:0.875rem;padding:0.65rem 1.4rem;border-radius:6px;text-decoration:none">
            View in Dashboard →
          </a>
        </div>
      </div>
    </div>
    """

    try:
        sg = SendGridAPIClient(api_key)
        for recipient_email, recipient_name in recipients:
            message = Mail(
                from_email=sender,
                to_emails=recipient_email,
                subject=subject,
                html_content=html_body
            )
            sg.send(message)
    except Exception:
        pass  # Never block a lead save due to email failure


def login_required(f):
    """Decorator that redirects to login if no active session."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not get_current_agent():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Decorator that requires is_admin = TRUE."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        agent = get_current_agent()
        if not agent:
            return redirect(url_for("login"))
        if not agent[4]:  # is_admin
            return jsonify({"status": "error", "message": "Admin access required."}), 403
        return f(*args, **kwargs)
    return decorated


# ─────────────────────────────────────────────
# TEST ROUTES
# ─────────────────────────────────────────────

@app.route("/ping")
def ping():
    return jsonify({"status": "ok", "message": "Flask is running."})


@app.route("/db-test")
def db_test():
    try:
        conn = get_connection()
        conn.close()
        return jsonify({"status": "ok", "message": "Database connection successful."})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/debug-agents-q7m2x")
def debug_agents():
    """Temporary read-only diagnostic — lists agents in whatever database
    this deployed app is actually connected to. No password hashes
    returned. Delete once the login issue is resolved."""
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT id, username, email, is_admin, is_active, created_at FROM agents ORDER BY id")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({
            "status": "ok",
            "count": len(rows),
            "agents": [
                {
                    "id": r[0], "username": r[1], "email": r[2],
                    "is_admin": r[3], "is_active": r[4],
                    "created_at": r[5].isoformat() if r[5] else None,
                }
                for r in rows
            ]
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/create-agent-w3fk8")
def create_agent_via_url():
    """Temporary one-time agent creation, reachable directly on the deployed
    app so it can't land in the wrong database. Query params: username,
    password (required); email, full_name, admin, notify (optional).
    Delete once the login issue is resolved."""
    username = (request.args.get("username") or "").strip().lower()
    password = (request.args.get("password") or "").strip()
    if not username or not password:
        return jsonify({"status": "error", "message": "username and password query params are required."}), 400

    full_name = request.args.get("full_name", "Josh Brown").strip()
    email     = (request.args.get("email") or f"{username}@centaurusenterprises.com").strip().lower()
    is_admin  = request.args.get("admin", "yes").lower() != "no"
    notify    = request.args.get("notify", "yes").lower() != "no"

    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT id FROM agents WHERE username = %s OR email = %s", (username, email))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"status": "error", "message": f"An agent with username '{username}' or email '{email}' already exists."}), 400

        salt    = secrets.token_hex(16)
        pw_hash = f"{salt}:{hashlib.sha256((salt + password).encode()).hexdigest()}"

        cur.execute("""
            INSERT INTO agents (full_name, email, username, password_hash, is_admin, notify_on_lead, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
        """, (full_name, email, username, pw_hash, is_admin, notify))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok", "message": f"Agent '{username}' created.", "username": username, "is_admin": is_admin})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/reset-agent-password-t9x4z")
def reset_agent_password():
    """Temporary one-time password reset, bypassing the duplicate-guard so
    it works even if the account already exists but the password is
    unknown. Query params: username, password (both required). Delete once
    the login issue is resolved."""
    username = (request.args.get("username") or "").strip().lower()
    password = (request.args.get("password") or "").strip()
    if not username or not password:
        return jsonify({"status": "error", "message": "username and password query params are required."}), 400

    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT id FROM agents WHERE username = %s", (username,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"status": "error", "message": f"No agent found with username '{username}'."}), 404

        salt    = secrets.token_hex(16)
        pw_hash = f"{salt}:{hashlib.sha256((salt + password).encode()).hexdigest()}"
        cur.execute("UPDATE agents SET password_hash = %s, is_active = TRUE WHERE username = %s", (pw_hash, username))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok", "message": f"Password reset for '{username}'."})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─────────────────────────────────────────────
# AUTH — LOGIN / LOGOUT
# ─────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    # Already logged in — go straight to dashboard
    if get_current_agent():
        return redirect(url_for("dashboard"))

    error = None

    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "").strip()

        try:
            conn = get_connection()
            cur  = conn.cursor()
            cur.execute(
                "SELECT id, full_name, password_hash, is_active FROM agents WHERE username = %s",
                (username,)
            )
            agent = cur.fetchone()
            cur.close()
            conn.close()

            if not agent:
                error = "Invalid username or password."
            elif not agent[3]:  # is_active
                error = "This account has been deactivated. Contact an admin."
            elif not verify_password(agent[2], password):
                error = "Invalid username or password."
            else:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(hours=8)
                session["agent_id"]   = agent[0]
                session["agent_name"] = agent[1]
                return redirect(url_for("dashboard"))

        except Exception as e:
            error = f"Database error: {str(e)}"

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─────────────────────────────────────────────
# DASHBOARD (placeholder — full UI coming next)
# ─────────────────────────────────────────────

@app.route("/admin")
@login_required
def dashboard():
    agent = get_current_agent()
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT id, submitted_at, product_type, first_name, last_name,
                   state, contact_preference, status
            FROM leads
            ORDER BY submitted_at DESC
        """)
        leads = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        leads = []

    return render_template("dashboard.html", agent=agent, leads=leads)


# ─────────────────────────────────────────────
# LEAD ROUTES
# ─────────────────────────────────────────────

@app.route("/lead/<int:lead_id>")
@login_required
def get_lead(lead_id):
    """Returns all fields for a single lead as JSON (for inline expansion)."""
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT * FROM leads WHERE id = %s", (lead_id,))
        row = cur.fetchone()
        cols = [desc[0] for desc in cur.description]
        cur.close()
        conn.close()

        if not row:
            return jsonify({"status": "error", "message": "Lead not found."}), 404

        lead = {}
        for i, col in enumerate(cols):
            val = row[i]
            if isinstance(val, datetime):
                val = val.strftime("%B %d, %Y at %I:%M %p")
            lead[col] = val

        return jsonify({"status": "ok", "lead": lead})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/lead/<int:lead_id>/status", methods=["POST"])
@login_required
def update_status(lead_id):
    data = request.get_json()

    valid_statuses = ["new", "follow_up", "does_not_pick_up", "sale", "no_sale"]
    new_status = data.get("status")
    if new_status and new_status not in valid_statuses:
        return jsonify({"status": "error", "message": "Invalid status value."}), 400

    def to_bool(val):
        if val is True or val == "true":  return True
        if val is False or val == "false": return False
        return None

    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("""
            UPDATE leads SET
                status                   = COALESCE(%s, status),
                notes                    = %s,
                first_name               = COALESCE(%s, first_name),
                last_name                = COALESCE(%s, last_name),
                age                      = %s,
                mobile_phone             = %s,
                home_phone               = %s,
                email                    = %s,
                city                     = %s,
                state                    = %s,
                product_type             = COALESCE(%s, product_type),
                coverage_amount          = %s,
                monthly_budget           = %s,
                has_beneficiary          = %s,
                beneficiary_relationship = %s,
                currently_insured        = %s,
                reason                   = %s,
                dob                      = %s,
                height_ft                = %s,
                height_in                = %s,
                weight                   = %s,
                tobacco                  = %s,
                major_conditions         = %s,
                medications              = %s,
                hobby                    = %s,
                assigned_agent           = %s
            WHERE id = %s
        """, (
            new_status,
            data.get("notes"),
            data.get("first_name"),
            data.get("last_name"),
            data.get("age"),
            data.get("mobile_phone"),
            data.get("home_phone"),
            data.get("email"),
            data.get("city"),
            data.get("state"),
            data.get("product_type"),
            data.get("coverage_amount"),
            data.get("monthly_budget"),
            data.get("has_beneficiary"),
            data.get("beneficiary_relationship"),
            to_bool(data.get("currently_insured")),
            data.get("reason"),
            data.get("dob") or None,
            data.get("height_ft"),
            data.get("height_in"),
            data.get("weight"),
            to_bool(data.get("tobacco")),
            data.get("major_conditions"),
            data.get("medications"),
            data.get("hobby"),
            data.get("assigned_agent"),
            lead_id
        ))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok"})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/lead/<int:lead_id>/delete", methods=["POST"])
@login_required
def delete_lead(lead_id):
    """Permanently deletes a lead from the database."""
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("DELETE FROM leads WHERE id = %s", (lead_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

# Display-friendly label maps for export
PRODUCT_LABELS = {
    'mortgage-protection': 'Mortgage Protection',
    'final-expense':       'Final Expense',
    'term-life':           'Term Life',
    'whole-life':          'Whole Life',
    'iul':                 'IUL / Wealth Building',
    'living-benefits':     'Living Benefits',
}

STATUS_LABELS = {
    'new':                'New',
    'follow_up':          'Follow Up',
    'does_not_pick_up':   'Does Not Pick Up',
    'sale':               'Sale',
    'no_sale':            'No Sale',
}

CONTACT_LABELS = {
    'call_me':           'Call Me',
    'book_appointment':  'Book Appointment',
}

def format_export_value(col, val):
    """Converts raw DB values to clean human-readable strings for export."""
    if val is None:
        return ''
    if col == 'submitted_at' and isinstance(val, datetime):
        return val.strftime('%b %d, %Y %I:%M %p')
    if col == 'product_type':
        return PRODUCT_LABELS.get(str(val), str(val))
    if col == 'status':
        return STATUS_LABELS.get(str(val), str(val))
    if col == 'contact_preference':
        return CONTACT_LABELS.get(str(val), str(val))
    return str(val)


@app.route("/export", methods=["POST"])
@login_required
def export_leads():
    """
    Exports selected leads as CSV or XLSX.
    Expects JSON: { "ids": [1,2,3], "columns": ["first_name","email",...], "format": "csv" }
    """
    import csv
    import io
    from flask import Response, send_file

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        openpyxl_available = True
    except ImportError:
        openpyxl_available = False

    data    = request.get_json()
    ids     = data.get("ids", [])
    columns = data.get("columns", [])
    fmt     = data.get("format", "csv")

    if not ids or not columns:
        return jsonify({"status": "error", "message": "No leads or columns selected."}), 400

    if fmt == "xlsx" and not openpyxl_available:
        return jsonify({"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}), 500

    # Whitelist columns to prevent SQL injection
    all_columns = [
        "id", "submitted_at", "product_type", "first_name", "last_name",
        "age", "email", "mobile_phone", "home_phone", "city", "state", "zip",
        "coverage_amount", "budget", "currently_insured", "beneficiary", "beneficiary_rel",
        "mp_lender", "mp_balance", "mp_monthly", "mp_years_remaining", "mp_purchase_year",
        "term_length", "term_reason", "term_annual_income", "wl_goal",
        "iul_annual_income", "iul_goal", "iul_investment_exp",
        "lb_concern", "lb_family_history",
        "tobacco", "height_ft", "height_in", "weight",
        "major_conditions", "minor_conditions", "medications",
        "contact_preference", "best_time", "hobby",
        "assigned_agent", "status", "notes"
    ]
    safe_columns = [c for c in columns if c in all_columns]

    if not safe_columns:
        return jsonify({"status": "error", "message": "No valid columns selected."}), 400

    col_str      = ", ".join(safe_columns)
    placeholders = ", ".join(["%s"] * len(ids))

    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute(
            f"SELECT {col_str} FROM leads WHERE id IN ({placeholders}) ORDER BY submitted_at DESC",
            ids
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

    # Format all values for clean export output
    formatted_rows = [
        [format_export_value(safe_columns[i], val) for i, val in enumerate(row)]
        for row in rows
    ]

    header_labels = [c.replace("_", " ").title() for c in safe_columns]

    if fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4B2E2B")

        for col_i, label in enumerate(header_labels, start=1):
            cell = ws.cell(row=1, column=col_i, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_i, row in enumerate(formatted_rows, start=2):
            for col_i, val in enumerate(row, start=1):
                ws.cell(row=row_i, column=col_i, value=val)

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="brown_agency_leads.xlsx"
        )

    else:
        # CSV with UTF-8 BOM so Excel opens it cleanly without encoding issues
        output = io.StringIO()
        output.write('\ufeff')  # UTF-8 BOM
        writer = csv.writer(output)
        writer.writerow(header_labels)
        writer.writerows(formatted_rows)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": "attachment; filename=brown_agency_leads.csv"}
        )


# ─────────────────────────────────────────────
# AGENT MANAGEMENT (admin only)
# ─────────────────────────────────────────────

@app.route("/agents")
@admin_required
def list_agents():
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT id, full_name, email, username, is_admin, notify_on_lead, is_active, created_at FROM agents ORDER BY created_at ASC")
        agents = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({"status": "ok", "agents": [
            {
                "id": a[0], "full_name": a[1], "email": a[2],
                "username": a[3], "is_admin": a[4],
                "notify_on_lead": a[5], "is_active": a[6],
                "created_at": a[7].strftime("%B %d, %Y") if a[7] else ""
            }
            for a in agents
        ]})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/agents/add", methods=["POST"])
@admin_required
def add_agent():
    import hashlib, secrets as sec
    data      = request.get_json()
    full_name = data.get("full_name", "").strip()
    email     = data.get("email", "").strip().lower()
    username  = data.get("username", "").strip().lower()
    password  = data.get("password", "").strip()
    is_admin  = data.get("is_admin", False)
    notify    = data.get("notify_on_lead", True)

    if not all([full_name, email, username, password]):
        return jsonify({"status": "error", "message": "All fields are required."}), 400

    salt      = sec.token_hex(16)
    pw_hash   = f"{salt}:{hashlib.sha256((salt + password).encode()).hexdigest()}"

    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT id FROM agents WHERE username = %s OR email = %s",
            (username, email)
        )
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"status": "error", "message": "Username or email already exists."}), 400

        cur.execute("""
            INSERT INTO agents (full_name, email, username, password_hash, is_admin, notify_on_lead, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
        """, (full_name, email, username, pw_hash, is_admin, notify))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok", "message": f"Agent {full_name} created."})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/agents/<int:agent_id>/toggle-notify", methods=["POST"])
@admin_required
def toggle_notify(agent_id):
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute(
            "UPDATE agents SET notify_on_lead = NOT notify_on_lead WHERE id = %s RETURNING notify_on_lead",
            (agent_id,)
        )
        result = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok", "notify_on_lead": result[0]})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/agents/<int:agent_id>/deactivate", methods=["POST"])
@admin_required
def deactivate_agent(agent_id):
    current = get_current_agent()
    if current and current[0] == agent_id:
        return jsonify({"status": "error", "message": "You cannot deactivate your own account."}), 400
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("UPDATE agents SET is_active = FALSE WHERE id = %s", (agent_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/agents/<int:agent_id>/reactivate", methods=["POST"])
@admin_required
def reactivate_agent(agent_id):
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("UPDATE agents SET is_active = TRUE WHERE id = %s", (agent_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─────────────────────────────────────────────
# SETTINGS PAGE
# ─────────────────────────────────────────────

@app.route("/agents/change-password", methods=["POST"])
@login_required
def change_password():
    import hashlib, secrets as sec
    data         = request.get_json()
    current_pw   = data.get("current_password", "")
    new_pw       = data.get("new_password", "")

    if not current_pw or not new_pw:
        return jsonify({"status": "error", "message": "Both fields are required."}), 400
    if len(new_pw) < 8:
        return jsonify({"status": "error", "message": "Password must be at least 8 characters."}), 400

    agent = get_current_agent()
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT password_hash FROM agents WHERE id = %s", (agent[0],))
        row = cur.fetchone()
        if not row or not verify_password(row[0], current_pw):
            cur.close(); conn.close()
            return jsonify({"status": "error", "message": "Current password is incorrect."}), 400

        salt    = sec.token_hex(16)
        pw_hash = f"{salt}:{hashlib.sha256((salt + new_pw).encode()).hexdigest()}"
        cur.execute("UPDATE agents SET password_hash = %s WHERE id = %s", (pw_hash, agent[0]))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/settings")
@admin_required
def settings():
    agent           = get_current_agent()
    mail_configured = bool(os.getenv("MAIL_SENDER") and os.getenv("MAIL_PASSWORD"))
    return render_template("settings.html", agent=agent, mail_configured=mail_configured)


# ─────────────────────────────────────────────
# AGENT APPLICATION (public)
# ─────────────────────────────────────────────

@app.route("/apply", methods=["POST"])
def apply():
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail

    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No data received."}), 400

        api_key = os.getenv("SENDGRID_API_KEY")
        sender  = os.getenv("MAIL_SENDER")

        if not api_key or not sender:
            return jsonify({"status": "ok"})

        first        = data.get("first_name", "")
        last         = data.get("last_name", "")
        email        = data.get("email", "")
        phone        = data.get("phone", "")
        state        = data.get("state", "")
        best_time    = data.get("best_time", "")
        licensed     = data.get("licensed", "")
        experience   = data.get("experience", "")
        arrangement  = data.get("arrangement", "")
        income_goal  = data.get("income_goal", "")
        interests    = ", ".join(data.get("interests", [])) or "None selected"
        message      = data.get("message", "")
        referral     = data.get("referral", "")
        submitted    = datetime.now().strftime("%B %d, %Y at %I:%M %p")

        subject = f"New Agent Application: {first} {last}"

        html_body = f"""
        <div style="font-family:'DM Sans',Arial,sans-serif;max-width:580px;margin:0 auto;background:#ffffff;border:1px solid #e8d5c0;border-radius:8px;overflow:hidden">
          <div style="background:#4B2E2B;padding:1.25rem 1.5rem">
            <p style="font-family:Arial,sans-serif;font-size:1.15rem;font-weight:800;color:#ffffff;margin:0">
              Brown<span style="color:#C08552">Financial Group</span>
            </p>
            <p style="font-size:0.75rem;color:#c8a882;margin:0.2rem 0 0;letter-spacing:0.08em;text-transform:uppercase">New Agent Application</p>
          </div>
          <div style="padding:1.5rem">
            <h2 style="font-size:1.1rem;color:#4B2E2B;margin:0 0 0.25rem">{first} {last}</h2>
            <p style="font-size:0.85rem;color:#8C5A3C;margin:0 0 1.5rem;font-weight:600">{state}</p>

            <table style="width:100%;border-collapse:collapse;font-size:0.875rem">
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;width:160px;font-weight:600">Email</td>
                <td style="padding:0.6rem 0;color:#2C1810">{email}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Phone</td>
                <td style="padding:0.6rem 0;color:#2C1810">{phone}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Best Time</td>
                <td style="padding:0.6rem 0;color:#2C1810">{best_time}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Licensed</td>
                <td style="padding:0.6rem 0;color:#2C1810">{licensed}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Experience</td>
                <td style="padding:0.6rem 0;color:#2C1810">{experience}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Arrangement</td>
                <td style="padding:0.6rem 0;color:#2C1810">{arrangement}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Income Goal</td>
                <td style="padding:0.6rem 0;color:#2C1810">{income_goal}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Interests</td>
                <td style="padding:0.6rem 0;color:#2C1810">{interests}</td>
              </tr>
              <tr style="border-bottom:1px solid #f0e4d4">
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Referral</td>
                <td style="padding:0.6rem 0;color:#2C1810">{referral}</td>
              </tr>
              <tr>
                <td style="padding:0.6rem 0;color:#9a7a6a;font-weight:600">Message</td>
                <td style="padding:0.6rem 0;color:#2C1810">{message if message else '—'}</td>
              </tr>
            </table>

            <p style="font-size:0.75rem;color:#c8a882;margin-top:1.25rem">Submitted {submitted}</p>
          </div>
        </div>
        """

        recipients = [
            "josh@thebrownfinancialgroup.com",
            "johnmbrown@outlook.com"
        ]

        sg = SendGridAPIClient(api_key)
        for recipient in recipients:
            message_obj = Mail(
                from_email=sender,
                to_emails=recipient,
                subject=subject,
                html_content=html_body
            )
            sg.send(message_obj)

        return jsonify({"status": "ok"})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



# ─────────────────────────────────────────────
# ONE-TIME MIGRATION (delete after use)
# ─────────────────────────────────────────────

@app.route("/run-migration-b3z7p1")
def run_migration_b():
    try:
        conn = get_connection()
        cur  = conn.cursor()
        for sql in [
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS gender VARCHAR(10)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS zip VARCHAR(20)",
        ]:
            cur.execute(sql)
        conn.commit()
        cur.close()
        conn.close()
        return "OK: gender and zip columns added", 200
    except Exception as e:
        return f"Error: {e}", 500


@app.route("/run-migration-mp9k2q")
def run_migration_mortgage_protection():
    """One-time migration adding columns needed by the protect-mortgage.com
    intake (/submit-mortgage-protection). Delete after use."""
    try:
        conn = get_connection()
        cur  = conn.cursor()
        for sql in [
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS code_word VARCHAR(20)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS code_word_set_at TIMESTAMPTZ",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS code_word_confirmed VARCHAR(20)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS code_word_confirmed_at TIMESTAMPTZ",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS homeowner VARCHAR(10)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS mortgage_balance VARCHAR(20)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS gclid VARCHAR(255)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS gbraid VARCHAR(255)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS wbraid VARCHAR(255)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS consent_version VARCHAR(10)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS consent_text TEXT",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS trustedform_cert_url VARCHAR(500)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS submitted_url TEXT",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS ip_address VARCHAR(64)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS user_agent VARCHAR(500)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS lead_source VARCHAR(100)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS lead_source_bucket VARCHAR(20)",
        ]:
            cur.execute(sql)
        conn.commit()
        cur.close()
        conn.close()
        return "OK: mortgage protection columns added", 200
    except Exception as e:
        return f"Error: {e}", 500


@app.route("/run-migration-init-h4v9t")
def run_migration_init():
    """One-time setup for a brand-new, empty Postgres instance. Creates
    leads and agents from scratch with every column referenced anywhere in
    this app (legacy multi-product form, export whitelist, dashboard, and
    the mortgage-protection funnel) so the two ALTER-TABLE migrations above
    aren't required to run first. Safe to re-run — CREATE TABLE IF NOT
    EXISTS never touches a table that already has data. Delete after use."""
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS leads (
                id SERIAL PRIMARY KEY,
                submitted_at TIMESTAMP DEFAULT NOW(),
                product_type VARCHAR(100),
                first_name VARCHAR(100),
                last_name VARCHAR(100),
                age INTEGER,
                gender VARCHAR(10),
                mobile_phone VARCHAR(50),
                home_phone VARCHAR(50),
                email VARCHAR(255),
                city VARCHAR(100),
                zip VARCHAR(20),
                state VARCHAR(100),
                dob DATE,
                tobacco BOOLEAN,
                has_beneficiary VARCHAR(10),
                beneficiary_relationship VARCHAR(100),
                beneficiary VARCHAR(10),
                beneficiary_rel VARCHAR(100),
                currently_insured BOOLEAN,
                coverage_amount VARCHAR(50),
                monthly_budget VARCHAR(50),
                budget VARCHAR(50),
                reason TEXT,
                height_ft INTEGER,
                height_in INTEGER,
                weight INTEGER,
                major_conditions TEXT,
                minor_conditions TEXT,
                medications TEXT,
                hobby VARCHAR(255),
                contact_preference VARCHAR(50),
                best_time VARCHAR(50),
                assigned_agent VARCHAR(100),
                status VARCHAR(50) DEFAULT 'new',
                notes TEXT,
                mp_lender VARCHAR(255),
                mp_balance VARCHAR(50),
                mp_monthly VARCHAR(50),
                mp_years_remaining VARCHAR(50),
                mp_purchase_year VARCHAR(10),
                term_length VARCHAR(50),
                term_reason TEXT,
                term_annual_income VARCHAR(50),
                wl_goal TEXT,
                iul_annual_income VARCHAR(50),
                iul_goal TEXT,
                iul_investment_exp TEXT,
                lb_concern TEXT,
                lb_family_history TEXT,
                code_word VARCHAR(20),
                code_word_set_at TIMESTAMPTZ,
                code_word_confirmed VARCHAR(20),
                code_word_confirmed_at TIMESTAMPTZ,
                homeowner VARCHAR(10),
                mortgage_balance VARCHAR(20),
                gclid VARCHAR(255),
                gbraid VARCHAR(255),
                wbraid VARCHAR(255),
                consent_version VARCHAR(10),
                consent_text TEXT,
                trustedform_cert_url VARCHAR(500),
                submitted_url TEXT,
                ip_address VARCHAR(64),
                user_agent VARCHAR(500),
                lead_source VARCHAR(100),
                lead_source_bucket VARCHAR(20)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS agents (
                id SERIAL PRIMARY KEY,
                full_name VARCHAR(255) NOT NULL,
                email VARCHAR(255) UNIQUE NOT NULL,
                username VARCHAR(100) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                is_admin BOOLEAN DEFAULT FALSE,
                notify_on_lead BOOLEAN DEFAULT TRUE,
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        return "OK: leads and agents tables created", 200
    except Exception as e:
        return f"Error: {e}", 500


# ─────────────────────────────────────────────
# FORM SUBMISSION (public)
# ─────────────────────────────────────────────

@app.route("/submit", methods=["POST"])
def submit():
    # -- CREATE TABLE leads (
    # --     id SERIAL PRIMARY KEY,
    # --     submitted_at TIMESTAMP DEFAULT NOW(),
    # --     product_type VARCHAR(100),
    # --     first_name VARCHAR(100),
    # --     last_name VARCHAR(100),
    # --     mobile_phone VARCHAR(50),
    # --     home_phone VARCHAR(50),
    # --     email VARCHAR(255),
    # --     city VARCHAR(100),
    # --     state VARCHAR(100),
    # --     has_beneficiary VARCHAR(10),
    # --     beneficiary_relationship VARCHAR(100),
    # --     reason TEXT,
    # --     contact_preference VARCHAR(50)
    # -- );
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No data received."}), 400

        conn = get_connection()
        cur  = conn.cursor()

        cur.execute("""
            INSERT INTO leads (
                product_type, first_name, last_name,
                mobile_phone, email, gender,
                city, zip, state, tobacco,
                has_beneficiary, beneficiary_relationship,
                reason, age, hobby
            ) VALUES (
                %(product_type)s, %(first_name)s, %(last_name)s,
                %(mobile_phone)s, %(email)s, %(gender)s,
                %(city)s, %(zip)s, %(state)s, %(tobacco)s,
                %(has_beneficiary)s, %(beneficiary_relationship)s,
                %(reason)s, %(age)s, %(hobby)s
            )
        """, {
            "product_type":             data.get("product_type"),
            "first_name":               data.get("first_name"),
            "last_name":                data.get("last_name"),
            "mobile_phone":             data.get("mobile_phone"),
            "email":                    data.get("email"),
            "gender":                   data.get("gender"),
            "city":                     data.get("city"),
            "zip":                      data.get("zip"),
            "state":                    data.get("state"),
            "tobacco":                  True if data.get("tobacco") == "yes" else (False if data.get("tobacco") == "no" else None),
            "has_beneficiary":          data.get("has_beneficiary"),
            "beneficiary_relationship": data.get("beneficiary_relationship"),
            "reason":                   data.get("reason"),
            "age":                      data.get("age"),
            "hobby":                    data.get("hobby"),
        })

        conn.commit()
        cur.close()
        conn.close()

        # Fire notification email to all agents with notify_on_lead = TRUE
        # Runs after DB commit so a mail failure never blocks the save
        send_lead_notification(data)

        return jsonify({"status": "ok", "message": "Lead saved successfully."})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─────────────────────────────────────────────
# MORTGAGE PROTECTION INTAKE (public) — protect-mortgage.com
# Browser never posts to the CRM directly; this is the server-side layer.
# Client-side validation is UX only — everything below is re-validated here.
# ─────────────────────────────────────────────

@app.route("/submit-mortgage-protection", methods=["POST"])
def submit_mortgage_protection():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"status": "error", "message": "No data received."}), 400

    required_fields = [
        "first_name", "last_name", "phone", "email", "zip",
        "date_of_birth", "mortgage_balance", "homeowner",
        "tobacco_use", "code_word",
    ]
    for field in required_fields:
        if not str(data.get(field, "")).strip():
            return jsonify({"status": "error", "field": field, "message": "This field is required."}), 400

    phone_digits = re.sub(r"\D", "", data.get("phone", ""))
    if len(phone_digits) not in (10, 11) or (len(phone_digits) == 11 and phone_digits[0] != "1"):
        return jsonify({"status": "error", "field": "phone", "message": "Enter a valid 10-digit US phone number."}), 400

    if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", data.get("email", "").strip()):
        return jsonify({"status": "error", "field": "email", "message": "Enter a valid email address."}), 400

    zip_code = data.get("zip", "").strip()
    if not re.fullmatch(r"\d{5}", zip_code):
        return jsonify({"status": "error", "field": "zip", "message": "Enter a valid 5-digit ZIP code."}), 400

    if data.get("homeowner") not in ("yes", "no"):
        return jsonify({"status": "error", "field": "homeowner", "message": "Please answer this question."}), 400

    if data.get("tobacco_use") not in ("yes", "no"):
        return jsonify({"status": "error", "field": "tobacco_use", "message": "Please answer this question."}), 400

    code_word_error = mp_validate_code_word(
        data.get("code_word", ""), data.get("first_name", ""), data.get("last_name", "")
    )
    if code_word_error:
        return jsonify({"status": "error", "field": "code_word", "message": code_word_error}), 400

    # Soft declines — not an error, not a lead. Checked after validation so a
    # malformed submission is still rejected rather than silently declined.
    if data.get("homeowner") == "no":
        return jsonify({
            "status": "declined",
            "message": "Thanks for your interest — mortgage protection coverage through this program requires homeownership. We're not able to help with this request right now."
        })

    if not mp_zip_allowed(zip_code):
        return jsonify({
            "status": "declined",
            "message": "Thanks for your interest — this program is not yet available in your area."
        })

    now = datetime.now(timezone.utc)

    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO leads (
                product_type, first_name, last_name, mobile_phone, email,
                zip, dob, tobacco, homeowner, mortgage_balance,
                code_word, code_word_set_at, code_word_confirmed,
                gclid, gbraid, wbraid,
                consent_version, consent_text, trustedform_cert_url,
                submitted_url, ip_address, user_agent,
                lead_source, lead_source_bucket
            ) VALUES (
                %(product_type)s, %(first_name)s, %(last_name)s, %(mobile_phone)s, %(email)s,
                %(zip)s, %(dob)s, %(tobacco)s, %(homeowner)s, %(mortgage_balance)s,
                %(code_word)s, %(code_word_set_at)s, %(code_word_confirmed)s,
                %(gclid)s, %(gbraid)s, %(wbraid)s,
                %(consent_version)s, %(consent_text)s, %(trustedform_cert_url)s,
                %(submitted_url)s, %(ip_address)s, %(user_agent)s,
                %(lead_source)s, %(lead_source_bucket)s
            )
        """, {
            "product_type":         "mortgage-protection",
            "first_name":           data.get("first_name", "").strip(),
            "last_name":            data.get("last_name", "").strip(),
            "mobile_phone":         "+1" + phone_digits if len(phone_digits) == 10 else "+" + phone_digits,
            "email":                data.get("email", "").strip(),
            "zip":                  zip_code,
            "dob":                  data.get("date_of_birth"),
            "tobacco":              data.get("tobacco_use") == "yes",
            "homeowner":            data.get("homeowner"),
            "mortgage_balance":     data.get("mortgage_balance"),
            "code_word":            data.get("code_word", "").strip(),
            "code_word_set_at":     now,
            "code_word_confirmed":  "pending",
            "gclid":                data.get("gclid") or None,
            "gbraid":               data.get("gbraid") or None,
            "wbraid":               data.get("wbraid") or None,
            "consent_version":      "1.0",
            "consent_text":         data.get("consent_text", ""),
            "trustedform_cert_url": data.get("trustedform_cert_url") or None,
            "submitted_url":        data.get("submitted_url", ""),
            "ip_address":           request.headers.get("X-Forwarded-For", request.remote_addr or ""),
            "user_agent":           request.headers.get("User-Agent", ""),
            "lead_source":          "protect-mortgage.com",
            "lead_source_bucket":   "approved",
        })

        conn.commit()
        cur.close()
        conn.close()

        send_lead_notification({
            "first_name":   data.get("first_name", "").strip(),
            "last_name":    data.get("last_name", "").strip(),
            "product_type": "mortgage-protection",
            "mobile_phone": data.get("phone", ""),
            "email":        data.get("email", ""),
            "state":        "",
            "city":         "",
        })

        return jsonify({"status": "ok"})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────


if __name__ == "__main__":
    app.run(port=5000, debug=True)